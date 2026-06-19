import pandas as pd
import numpy as np
import re
import os

# Resolve paths relative to this script's directory for Streamlit Cloud compatibility
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
hino_prices_path = os.path.join(BASE_DIR, "INCREMENTO DE PRECIOS HINO.xlsx")
historical_data_path = os.path.join(BASE_DIR, "TD MAE TEOJAMA HISTORICO.xlsx")
hino_costs_path = os.path.join(BASE_DIR, "COSTOS.xlsx")

def clean_model_name(name):
    """Clean model name for matching between prices and sales datasets."""
    name = str(name).strip().upper()
    # Fix the common typo XZlJ650L / XZLJ650L -> XZU650L
    name = re.sub(r'XZ[LIJ]+650', 'XZU650', name)
    name = name.replace(" ", "")
    return name

def load_hino_prices():
    """Load and format the HINO prices sheet, resolving simple formulas like =+G<row>."""
    import openpyxl
    
    wb = openpyxl.load_workbook(hino_prices_path, data_only=False)
    ws = wb["PRICES"]
    
    months = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUN"]
    
    records = []
    # Rows 5 to 20 contain the 16 Hino models in 1-based indexing
    for r in range(5, 21):
        model = ws.cell(row=r, column=3).value
        prices_dict = {"MODEL_RAW": model, "MODEL_CLEAN": clean_model_name(model)}
        
        for i, month in enumerate(months):
            col = 4 + i  # January is column D (4)
            val = ws.cell(row=r, column=col).value
            
            # Resolve formula if it is a string starting with '='
            if isinstance(val, str) and val.startswith("="):
                formula = val.replace("=", "").replace("+", "").strip().upper()
                if len(formula) >= 2 and formula[0] == 'G':
                    try:
                        ref_row = int(formula[1:])
                        # G is column 7
                        val = float(ws.cell(row=ref_row, column=7).value)
                    except:
                        pass
            
            prices_dict[month] = float(val) if val is not None else 0.0
            
        records.append(prices_dict)
        
    df = pd.DataFrame(records)
    
    # Map each model to its segment and series based on the official list
    def assign_seg_series(model_clean):
        m = model_clean
        if "XZU640" in m:
            return "LIGHT", "SERIE 300"
        elif "XZU650" in m or "XZL" in m or "XZI" in m:
            return "LIGHT", "SERIE 300"
        elif "XZU710" in m:
            return "LIGHT", "SERIE 300"
        elif "XZU720" in m:
            return "LIGHT", "SERIE 300"
        elif "FC9" in m:
            return "MEDIUM", "SERIE 500"
        elif "GD8" in m:
            return "MEDIUM", "SERIE 500"
        elif "GH8" in m:
            return "MEDIUM", "SERIE 500"
        elif "FM2" in m:
            return "MEDIUM", "SERIE 500"
        elif "AK8" in m:
            return "BUS", "BUS"
        elif "RM1" in m:
            return "BUS", "BUS"
        elif "SS1" in m:
            return "TRACTO", "SERIE 700"
        elif "FS1" in m:
            return "HEAVY", "SERIE 700"
        else:
            return "OTROS", "OTROS"
            
    res = df['MODEL_CLEAN'].apply(assign_seg_series)
    df['SEGMENTO'] = [r[0] for r in res]
    df['SERIES'] = [r[1] for r in res]
    
    # Exclude Buses from pricing sheet
    df = df[df['SEGMENTO'] != 'BUS']
    return df

def load_hino_costs():
    """Load and format the HINO costs sheet from COSTOS.xlsx."""
    df_costs = pd.read_excel(hino_costs_path, sheet_name="Hoja1", header=1)
    df_costs['MODEL_CLEAN'] = df_costs['MODELO'].apply(clean_model_name)
    
    # Translate month columns to standard English for consistency
    month_translation = {
        "ENERO": "JANUARY",
        "FEBRERO": "FEBRUARY",
        "MARZO": "MARCH",
        "ABRIL": "APRIL",
        "MAYO": "MAY",
        "JUNIO": "JUN"
    }
    df_costs = df_costs.rename(columns=month_translation)
    return df_costs

def load_sales_data():
    """Load and parse historical sales data from AEADE."""
    df_tga = pd.read_excel(historical_data_path, sheet_name="TD GENERAL A", header=None)
    
    # Parse headers by combining Year (Row 9) and Month (Row 10)
    new_headers = []
    for col in range(df_tga.shape[1]):
        val_r9 = df_tga.iloc[9, col]
        val_r10 = df_tga.iloc[10, col]
        
        if col < 10:
            new_headers.append(val_r10 if pd.notna(val_r10) else f"Col_{col}")
        else:
            try:
                year = int(float(val_r9))
                month = int(float(val_r10))
                new_headers.append(f"{year}_{month:02d}")
            except:
                if pd.notna(val_r9):
                    new_headers.append(str(val_r9).strip().replace(" ", "_"))
                else:
                    new_headers.append(f"Col_{col}")
                    
    df_sales = df_tga.iloc[11:].copy()
    df_sales.columns = new_headers
    
    # Filter out BUS segment
    df_sales = df_sales[df_sales['SEGMENTO'] != 'BUS']
    
    # Filter HINO, CHEVROLET, and FUSO raw rows (exclude totals)
    df_hino = df_sales[df_sales['MARCA'] == 'HINO'].copy()
    df_chev = df_sales[df_sales['MARCA'] == 'CHEVROLET'].copy()
    df_fuso = df_sales[df_sales['MARCA'] == 'FUSO'].copy()
    
    # Months in 2026 available in the dataset
    months_2026 = ["2026_01", "2026_02", "2026_03", "2026_04", "2026_05"]
    for col in months_2026:
        df_hino[col] = pd.to_numeric(df_hino[col], errors='coerce').fillna(0)
        df_chev[col] = pd.to_numeric(df_chev[col], errors='coerce').fillna(0)
        df_fuso[col] = pd.to_numeric(df_fuso[col], errors='coerce').fillna(0)
        
    return df_hino, df_chev, df_fuso, months_2026

def calculate_simulation(elasticity_light=-4.6, elasticity_other=-4.6, shift_factor=0.5, hino_june_sales=235,
                         sim_model_type="Price Elasticity", hino_share_2025=5.4, market_growth_rate=35.9,
                         start_at_isuzu_baseline=True):
    """Run simulation logic for Ene-Jun 2026."""
    df_prices = load_hino_prices()
    df_hino_sales, df_chev_sales, df_fuso_sales, months_db = load_sales_data()
    
    months_map = {
        "2026_01": "JANUARY",
        "2026_02": "FEBRUARY",
        "2026_03": "MARCH",
        "2026_04": "APRIL",
        "2026_05": "MAY",
        "2026_06": "JUN"
    }
    
    # 1. Process Hino Portfolio Sales for Jan-May 2026
    portfolio_records = []
    for _, p_row in df_prices.iterrows():
        model_clean = p_row['MODEL_CLEAN']
        model_raw = p_row['MODEL_RAW']
        segment = p_row['SEGMENTO']
        series = p_row['SERIES']
        
        # Find matching sales rows in AEADE
        matching_rows = df_hino_sales[df_hino_sales['MODELO'].apply(lambda x: model_clean in clean_model_name(x))]
        
        model_sales = {}
        for m_db in months_db:
            model_sales[m_db] = matching_rows[m_db].sum()
            
        rec = {
            "MODEL_RAW": model_raw,
            "MODEL_CLEAN": model_clean,
            "SEGMENTO": segment,
            "SERIES": series,
            **model_sales
        }
        portfolio_records.append(rec)
        
    df_portfolio_sales = pd.DataFrame(portfolio_records)
    
    # 2. Calculate "Otros Modelos Hino" to square with AEADE total (865 u Ene-May)
    hino_aeade_totals = {m: df_hino_sales[m].sum() for m in months_db}
    portfolio_totals_by_month = {m: df_portfolio_sales[m].sum() for m in months_db}
    other_sales_by_month = {m: hino_aeade_totals[m] - portfolio_totals_by_month[m] for m in months_db}
    
    other_row = {
        "MODEL_RAW": "Other Hino Models",
        "MODEL_CLEAN": "OTROS_MODELOS_HINO",
        "SEGMENTO": "OTROS",
        "SERIES": "OTROS",
        **other_sales_by_month
    }
    df_portfolio_sales = pd.concat([df_portfolio_sales, pd.DataFrame([other_row])], ignore_index=True)
    
    # 3. Add June sales (dynamic Hino sales projected by user, distributed based on Ene-May share)
    db_months_list = ["2026_01", "2026_02", "2026_03", "2026_04", "2026_05"]
    df_portfolio_sales['Total_Ene_May'] = df_portfolio_sales[db_months_list].sum(axis=1)
    grand_total_ene_may = df_portfolio_sales['Total_Ene_May'].sum()
    
    # Distribute dynamic June units
    june_sales = []
    for idx, row in df_portfolio_sales.iterrows():
        share = row['Total_Ene_May'] / grand_total_ene_may if grand_total_ene_may > 0 else 0
        june_sales.append(round(share * hino_june_sales))
        
    # Correct rounding error so that sum is exactly hino_june_sales
    diff = hino_june_sales - sum(june_sales)
    if diff != 0:
        largest_idx = np.argmax(df_portfolio_sales['Total_Ene_May'].values)
        june_sales[largest_idx] += diff
        
    df_portfolio_sales["2026_06"] = june_sales
    df_portfolio_sales.drop(columns=['Total_Ene_May'], inplace=True)
    
    # Add Others price row to df_prices (assumed average prices, variation = 0%)
    avg_prices = {}
    for month_name in ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUN"]:
        avg_prices[month_name] = df_prices[month_name].mean()
        
    other_price_row = {
        "MODEL_RAW": "Other Hino Models",
        "MODEL_CLEAN": "OTROS_MODELOS_HINO",
        "SEGMENTO": "OTROS",
        "SERIES": "OTROS",
        **avg_prices
    }
    df_prices = pd.concat([df_prices, pd.DataFrame([other_price_row])], ignore_index=True)
    
    # 4. Simulate counterfactual Hino sales model by model
    all_months = ["2026_01", "2026_02", "2026_03", "2026_04", "2026_05", "2026_06"]
    
    if sim_model_type == "2025 Market Share & Growth":
        # Load total market from AEADE database to compute total market size by month
        df_tga_raw = pd.read_excel(historical_data_path, sheet_name="TD GENERAL A", header=None)
        # Parse headers
        new_headers = []
        for col in range(df_tga_raw.shape[1]):
            val_r9 = df_tga_raw.iloc[9, col]
            val_r10 = df_tga_raw.iloc[10, col]
            if col < 10:
                new_headers.append(val_r10 if pd.notna(val_r10) else f"Col_{col}")
            else:
                try:
                    year = int(float(val_r9))
                    month = int(float(val_r10))
                    new_headers.append(f"{year}_{month:02d}")
                except:
                    if pd.notna(val_r9):
                        new_headers.append(str(val_r9).strip().replace(" ", "_"))
                    else:
                        new_headers.append(f"Col_{col}")
        df_sales_raw = df_tga_raw.iloc[11:].copy()
        df_sales_raw.columns = new_headers
        
        # Clean columns
        months_needed = ["2025_01", "2025_02", "2025_03", "2025_04", "2025_05", "2025_06",
                         "2026_01", "2026_02", "2026_03", "2026_04", "2026_05"]
        for m in months_needed:
            df_sales_raw[m] = pd.to_numeric(df_sales_raw[m], errors='coerce').fillna(0)
            
        market_by_month_raw = {m: df_sales_raw[m].sum() for m in months_needed}
        
        # Estimate total market for June 2026 based on June 2025 and growth rate
        market_by_month_raw["2026_06"] = market_by_month_raw["2025_06"] * (1 + market_growth_rate / 100.0)
        
        # Calculate model weights based on Jan-May 2025 Hino sales
        months_2025 = ["2025_01", "2025_02", "2025_03", "2025_04", "2025_05"]
        hino_sales_2025_sum = df_hino_sales[months_2025].sum().sum()
        
        portfolio_weights = {}
        for _, p_row in df_prices.iterrows():
            model_clean = p_row['MODEL_CLEAN']
            matching_rows = df_hino_sales[df_hino_sales['MODELO'].apply(lambda x: model_clean in clean_model_name(x))]
            model_sales_2025 = matching_rows[months_2025].sum().sum()
            weight = model_sales_2025 / hino_sales_2025_sum if hino_sales_2025_sum > 0 else 0
            portfolio_weights[model_clean] = weight
            
        # For other models Hino:
        portfolio_weights["OTROS_MODELOS_HINO"] = 1.0 - sum(portfolio_weights.values())
        
    simulated_records = []
    for _, row in df_portfolio_sales.iterrows():
        model_clean = row['MODEL_CLEAN']
        model_raw = row['MODEL_RAW']
        segment = row['SEGMENTO']
        
        # Get elasticity
        if segment == "LIGHT":
            eps = elasticity_light
        elif segment == "OTROS":
            eps = 0.0 # No arancel impact on other minor Hino models
        else:
            eps = elasticity_other
            
        p_row = df_prices[df_prices['MODEL_CLEAN'] == model_clean].iloc[0]
        p_jan = p_row['JANUARY']
        
        res = {
            "MODEL_RAW": model_raw,
            "MODEL_CLEAN": model_clean,
            "SEGMENTO": segment,
            "SERIES": row['SERIES']
        }
        
        for m_db in all_months:
            p_month_name = months_map[m_db]
            p_act = p_row[p_month_name]
            
            # If segment is OTROS, we assume simulated price = real price (no arancel impact)
            if segment == "OTROS":
                p_sim = p_act
            else:
                p_sim = p_jan
                
            pct_price_inc = (p_act - p_sim) / p_sim if p_sim > 0 else 0
            sales_act = row[m_db]
            
            if sim_model_type == "2025 Market Share & Growth":
                # Simulated Hino sales under market share & growth model
                hino_total_sim_m = market_by_month_raw[m_db] * (hino_share_2025 / 100.0)
                sales_sim = hino_total_sim_m * portfolio_weights.get(model_clean, 0.0)
            else:
                # Price elasticity model
                sales_sim = sales_act * (1 - (eps * pct_price_inc))
                
            sales_sim = max(0.0, sales_sim)
            
            res[f"{m_db}_PRECIO_REAL"] = p_act
            res[f"{m_db}_PRECIO_SIM"] = p_sim
            res[f"{m_db}_PRECIO_VAR"] = pct_price_inc
            res[f"{m_db}_VENTAS_REAL"] = sales_act
            res[f"{m_db}_VENTAS_SIM"] = round(sales_sim)
            res[f"{m_db}_VENTAS_PERDIDAS"] = round(sales_sim) - sales_act
            
        simulated_records.append(res)
        
    df_hino_sim = pd.DataFrame(simulated_records)
    
    # 5. Process Chevrolet and Fuso sales
    chev_totals = {}
    fuso_totals = {}
    for m_db in months_db:
        chev_totals[m_db] = df_chev_sales[m_db].sum()
        fuso_totals[m_db] = df_fuso_sales[m_db].sum()
        
    # Project June for Chevrolet and Fuso as their Jan-May averages
    chev_totals["2026_06"] = round(sum(chev_totals.values()) / len(months_db))
    fuso_totals["2026_06"] = round(sum(fuso_totals.values()) / len(months_db))
    
    # Segment breakdown for Chevrolet (needed for competitor segment shifts)
    chev_segment_records = []
    for m_db in all_months:
        rec = {"MES": m_db, "TOTAL": chev_totals[m_db]}
        if m_db != "2026_06":
            rec["LIGHT"] = df_chev_sales[df_chev_sales['SEGMENTO'] == 'LIGHT'][m_db].sum()
            rec["MEDIUM"] = df_chev_sales[df_chev_sales['SEGMENTO'] == 'MEDIUM'][m_db].sum()
            rec["HEAVY_TRACTO"] = df_chev_sales[df_chev_sales['SEGMENTO'].isin(['HEAVY', 'TRACTO'])][m_db].sum()
        else:
            light_share = sum(df_chev_sales[df_chev_sales['SEGMENTO'] == 'LIGHT'][months_db].sum()) / sum(chev_totals[m] for m in months_db)
            med_share = sum(df_chev_sales[df_chev_sales['SEGMENTO'] == 'MEDIUM'][months_db].sum()) / sum(chev_totals[m] for m in months_db)
            rec["LIGHT"] = round(light_share * chev_totals[m_db])
            rec["MEDIUM"] = round(med_share * chev_totals[m_db])
            rec["HEAVY_TRACTO"] = chev_totals[m_db] - rec["LIGHT"] - rec["MEDIUM"]
        chev_segment_records.append(rec)
    df_chev_monthly = pd.DataFrame(chev_segment_records).set_index("MES")
    
    if start_at_isuzu_baseline:
        simulated_jan_total = df_hino_sim["2026_01_VENTAS_SIM"].sum()
        isuzu_jan_total = chev_totals["2026_01"]
        scale_factor = isuzu_jan_total / simulated_jan_total if simulated_jan_total > 0 else 1.0
        
        for m_db in all_months:
            # Scale simulated sales and round to integer
            df_hino_sim[f"{m_db}_VENTAS_SIM"] = (df_hino_sim[f"{m_db}_VENTAS_SIM"] * scale_factor).round().astype(int)
            df_hino_sim[f"{m_db}_VENTAS_PERDIDAS"] = df_hino_sim[f"{m_db}_VENTAS_SIM"] - df_hino_sim[f"{m_db}_VENTAS_REAL"]
            
    # 6. Apply Chevrolet Sales Shift
    chev_sim_records = []
    for m_db in all_months:
        hino_m_real = df_hino_sim[f"{m_db}_VENTAS_REAL"].sum()
        hino_m_sim = df_hino_sim[f"{m_db}_VENTAS_SIM"].sum()
        hino_m_lost = hino_m_sim - hino_m_real
        
        chev_m_real = chev_totals[m_db]
        chev_m_sim = chev_m_real - (shift_factor * hino_m_lost)
        
        chev_light_real = df_chev_monthly.loc[m_db, "LIGHT"]
        chev_medium_real = df_chev_monthly.loc[m_db, "MEDIUM"]
        
        hino_light_lost = df_hino_sim[df_hino_sim['SEGMENTO'] == 'LIGHT'][f"{m_db}_VENTAS_PERDIDAS"].sum()
        hino_medium_lost = df_hino_sim[df_hino_sim['SEGMENTO'] == 'MEDIUM'][f"{m_db}_VENTAS_PERDIDAS"].sum()
        
        chev_light_sim = chev_light_real - (shift_factor * hino_light_lost)
        chev_medium_sim = chev_medium_real - (shift_factor * hino_medium_lost)
        
        chev_sim_records.append({
            "MES": m_db,
            "CHEV_TOTAL_REAL": chev_m_real,
            "CHEV_TOTAL_SIM": round(chev_m_sim),
            "CHEV_LIGHT_REAL": chev_light_real,
            "CHEV_LIGHT_SIM": round(chev_light_sim),
            "CHEV_MEDIUM_REAL": chev_medium_real,
            "CHEV_MEDIUM_SIM": round(chev_medium_sim)
        })
        
    df_chev_sim = pd.DataFrame(chev_sim_records).set_index("MES")
    
    # Convert Fuso to a simple DataFrame for easy dashboard access
    df_fuso_monthly = pd.DataFrame([
        {"MES": m, "FUSO_TOTAL_REAL": fuso_totals[m]} for m in all_months
    ]).set_index("MES")
    
    return df_hino_sim, df_chev_sim, df_fuso_monthly, df_prices

def load_total_market_data():
    """Load total market sales data (excluding summary rows) from AEADE."""
    df_tga = pd.read_excel(historical_data_path, sheet_name="TD GENERAL A", header=None)
    
    # Parse headers by combining Year (Row 9) and Month (Row 10)
    new_headers = []
    for col in range(df_tga.shape[1]):
        val_r9 = df_tga.iloc[9, col]
        val_r10 = df_tga.iloc[10, col]
        
        if col < 10:
            new_headers.append(val_r10 if pd.notna(val_r10) else f"Col_{col}")
        else:
            try:
                year = int(float(val_r9))
                month = int(float(val_r10))
                new_headers.append(f"{year}_{month:02d}")
            except:
                if pd.notna(val_r9):
                    new_headers.append(str(val_r9).strip().replace(" ", "_"))
                else:
                    new_headers.append(f"Col_{col}")
                    
    df_sales = df_tga.iloc[11:].copy()
    df_sales.columns = new_headers
    df_sales = df_sales[df_sales['SEGMENTO'] != 'BUS']
    
    # Filter out summary rows (double counting)
    df_brands_clean = df_sales[~df_sales['MARCA'].str.startswith('Total', na=False)].copy()
    
    # Convert columns to numeric
    months_2026 = ["2026_01", "2026_02", "2026_03", "2026_04", "2026_05"]
    for col in months_2026:
        df_brands_clean[col] = pd.to_numeric(df_brands_clean[col], errors='coerce').fillna(0)
        
    return df_brands_clean

def forecast_june_holt_winters():
    """Forecast June 2026 sales volume using Seasonal Holt-Winters model."""
    df_hino, _, _, _ = load_sales_data()
    all_cols = list(df_hino.columns)
    monthly_cols = [c for c in all_cols if len(c) == 7 and c[4] == '_' and c[:4].isdigit() and c[5:].isdigit()]
    monthly_cols.sort()
    
    hino_monthly = df_hino[monthly_cols].sum(axis=0).astype(float)
    dates = pd.to_datetime([c.replace('_', '-') for c in monthly_cols], format='%Y-%m')
    hino_monthly.index = pd.DatetimeIndex(dates, freq='MS')
    
    from statsmodels.tsa.holtwinters import ExponentialSmoothing
    model = ExponentialSmoothing(hino_monthly, trend='add', seasonal='mul', seasonal_periods=12)
    fit = model.fit()
    pred = fit.forecast(1)
    return int(round(pred.iloc[0]))


if __name__ == "__main__":
    print("Testing Simulation Logic...")
    hino_sim, chev_sim, fuso_sim, prices = calculate_simulation()
    print("\nSimulation completed successfully!")
