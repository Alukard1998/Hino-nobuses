import openpyxl
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import os
import sys

# Add directory to path to import local module
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from simulador_logica import calculate_simulation, load_hino_costs, forecast_june_holt_winters

# Colors
C_DARK_BLUE   = "1A3A5C"   # Teojama / Hino Navy
C_HINO_RED    = "C0392B"   # Hino Brand Red
C_GOLD        = "F0B429"   # Accent Gold
C_LIGHT_BLUE  = "D6E4F0"   # Light Fill
C_LIGHT_GRAY  = "F2F2F2"
C_WHITE       = "FFFFFF"
C_MUTED       = "7F8C8D"
C_GREEN       = "27AE60"

def thin_border():
    s = Side(style='thin', color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def data_cell(ws, cell, value, fill=None, bold=False, align="center", color="000000", size=9, num_format=None):
    ws[cell] = value
    ws[cell].font = Font(name="Arial", size=size, bold=bold, color=color)
    if fill:
        ws[cell].fill = header_fill(fill)
    ws[cell].alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    ws[cell].border = thin_border()
    if num_format:
        ws[cell].number_format = num_format

def merge_title(ws, cell, text, size=14, bold=True, color=C_WHITE, fill=C_DARK_BLUE, align="center"):
    ws[cell] = text
    ws[cell].font = Font(name="Arial", size=size, bold=bold, color=color)
    ws[cell].fill = header_fill(fill)
    ws[cell].alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def generate_excel_report(elasticity_light=-4.6, elasticity_heavy=-4.6, shift_factor=0.5, june_sales_val=None,
                          sim_model_type="Price Elasticity", hino_share_2025=5.4, market_growth_rate=35.9,
                          start_at_isuzu_baseline=True):
    if june_sales_val is None:
        try:
            june_sales_val = forecast_june_holt_winters()
        except Exception as e:
            june_sales_val = 196
    df_hino_sim, df_chev_sim, _, df_pr = calculate_simulation(
        elasticity_light=elasticity_light,
        elasticity_other=elasticity_heavy,
        shift_factor=shift_factor,
        hino_june_sales=june_sales_val,
        sim_model_type=sim_model_type,
        hino_share_2025=hino_share_2025,
        market_growth_rate=market_growth_rate,
        start_at_isuzu_baseline=start_at_isuzu_baseline
    )
    
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # TAB 1: RESUMEN EJECUTIVO
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "1. Executive Summary"
    ws1.sheet_view.showGridLines = True
    
    # Set column widths
    col_widths = {
        'A': 18, 'B': 12, 'C': 15, 'D': 15, 'E': 12, 
        'F': 18, 'G': 18, 'H': 18, 'I': 15, 'J': 15, 
        'K': 15, 'L': 15
    }
    for col, w in col_widths.items():
        ws1.column_dimensions[col].width = w
        
    # Title Block
    ws1.merge_cells("A1:L1")
    merge_title(ws1, "A1", "COLOMBIA TARIFF IMPACT — HINO PRICE & SALES ANALYSIS VS. ISUZU", size=13, bold=True)
    ws1.row_dimensions[1].height = 40
    
    ws1.merge_cells("A2:L2")
    ws1["A2"] = f"Market Research Department — Teojama Comercial S.A. | Simulation Report | Updated: {datetime.date.today().strftime('%d/%m/%Y')} | Source: AEADE & Teojama Pricing"
    ws1["A2"].font = Font(name="Arial", size=8, italic=True, color=C_MUTED)
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1["A2"].fill = header_fill("EBF5FB")
    ws1.row_dimensions[2].height = 20
    
    # Section 1: KPIs
    ws1.merge_cells("A4:L4")
    merge_title(ws1, "A4", "GENERAL IMPACT METRICS (PERIOD JANUARY - JUNE 2026)", size=10, bold=True, fill=C_HINO_RED)
    ws1.row_dimensions[4].height = 22
    
    # Aggregate stats
    all_months = ["2026_01", "2026_02", "2026_03", "2026_04", "2026_05", "2026_06"]
    
    hino_total_real = sum(df_hino_sim[f"{m}_VENTAS_REAL"].sum() for m in all_months)
    hino_total_sim = sum(df_hino_sim[f"{m}_VENTAS_SIM"].sum() for m in all_months)
    hino_total_lost = hino_total_sim - hino_total_real
    
    rev_total_real = sum((df_hino_sim[f"{m}_VENTAS_REAL"] * df_hino_sim[f"{m}_PRECIO_REAL"]).sum() for m in all_months)
    rev_total_sim = sum((df_hino_sim[f"{m}_VENTAS_SIM"] * df_hino_sim[f"{m}_PRECIO_SIM"]).sum() for m in all_months)
    rev_total_lost = rev_total_sim - rev_total_real
    
    chev_total_real = sum(df_chev_sim.loc[m, "CHEV_TOTAL_REAL"] for m in all_months)
    chev_total_sim = sum(df_chev_sim.loc[m, "CHEV_TOTAL_SIM"] for m in all_months)
    
    kpis = [
        ("Hino Real Sales (units)", hino_total_real, "#,##0", "Hino registration volume recorded in AEADE (with tariff)"),
        ("Hino Simulated Sales (units)", hino_total_sim, "#,##0", "Hino estimated volume if the tariff had remained at 0%"),
        ("Lost Sales Volume (units)", hino_total_lost, "#,##0", "Volume of unsold trucks due to list-price increases"),
        ("Estimated Real Revenue (USD)", rev_total_real, "$#,##0", "Estimated gross revenue from list prices of analyzed models"),
        ("Estimated Simulated Revenue (USD)", rev_total_sim, "$#,##0", "Projected gross revenue under the counterfactual 0% tariff scenario"),
        ("Gross Revenue Loss (USD)", rev_total_lost, "$#,##0", "Estimated gross revenue loss from reduced sales volume"),
    ]
    
    # Write KPIs
    for i, (kpi_name, kpi_val, kpi_fmt, kpi_desc) in enumerate(kpis):
        row = 5 + i
        ws1.row_dimensions[row].height = 20
        ws1.merge_cells(f"A{row}:C{row}")
        data_cell(ws1, f"A{row}", kpi_name, fill=C_DARK_BLUE, bold=True, color=C_WHITE, align="left")
        ws1.merge_cells(f"D{row}:E{row}")
        
        # Color formatting for visual impact
        val_color = C_HINO_RED if "Lost" in kpi_name or "Loss" in kpi_name else "000000"
        bold_val = True if "Lost" in kpi_name or "Loss" in kpi_name else False
        fill_val = "FDEDEC" if "Lost" in kpi_name or "Loss" in kpi_name else None
        
        data_cell(ws1, f"D{row}", kpi_val, fill=fill_val, bold=bold_val, color=val_color, align="right", num_format=kpi_fmt)
        ws1.merge_cells(f"F{row}:L{row}")
        data_cell(ws1, f"F{row}", kpi_desc, align="left", size=8.5, color="555555")
        
    # Section 2: Monthly Breakdown
    start_breakdown_row = 13
    ws1.merge_cells(f"A{start_breakdown_row}:L{start_breakdown_row}")
    merge_title(ws1, f"A{start_breakdown_row}", "MONTHLY PRICE & SALES EVOLUTION (REAL VS. SIMULATED)", size=10, bold=True, fill=C_HINO_RED)
    ws1.row_dimensions[start_breakdown_row].height = 22
    
    # Headers for breakdown table
    breakdown_headers = [
        "Month", "Colombia Tariff", 
        "Avg. Real Price", "Avg. Simulated Price", "Price Inc. %",
        "Hino Real Sales", "Hino Simulated Sales", "Lost Sales Volume",
        "Isuzu Real Sales", "Isuzu Sim Sales", "Hino Market Share (Real)", "Hino Market Share (Sim)"
    ]
    
    header_row = start_breakdown_row + 1
    ws1.row_dimensions[header_row].height = 25
    for col_idx, text in enumerate(breakdown_headers):
        cell_ref = f"{get_column_letter(col_idx + 1)}{header_row}"
        data_cell(ws1, cell_ref, text, fill=C_DARK_BLUE, bold=True, color=C_WHITE, size=8.5)
        
    # Monthly data
    months_names = {
        "2026_01": ("January", 0.0),
        "2026_02": ("February", 0.3),
        "2026_03": ("March", 0.5),
        "2026_04": ("April", 0.5),
        "2026_05": ("May", 1.0),
        "2026_06": ("June (Est.)", 0.0)
    }
    
    data_idx = 0
    for m_db, (m_lbl, tariff) in months_names.items():
        row = header_row + 1 + data_idx
        ws1.row_dimensions[row].height = 20
        
        # Calculate values for Hino
        h_real = df_hino_sim[f"{m_db}_VENTAS_REAL"].sum()
        h_sim = df_hino_sim[f"{m_db}_VENTAS_SIM"].sum()
        h_lost = df_hino_sim[f"{m_db}_VENTAS_PERDIDAS"].sum()
        
        p_avg_real = (df_hino_sim[f"{m_db}_PRECIO_REAL"] * df_hino_sim[f"{m_db}_VENTAS_REAL"]).sum() / h_real if h_real > 0 else 0
        p_avg_sim = (df_hino_sim[f"{m_db}_PRECIO_SIM"] * df_hino_sim[f"{m_db}_VENTAS_SIM"]).sum() / h_sim if h_sim > 0 else 0
        p_var = (p_avg_real - p_avg_sim) / p_avg_sim if p_avg_sim > 0 else 0
        
        c_real = df_chev_sim.loc[m_db, "CHEV_TOTAL_REAL"]
        c_sim = df_chev_sim.loc[m_db, "CHEV_TOTAL_SIM"]
        
        ms_real = h_real / (h_real + c_real) if (h_real + c_real) > 0 else 0
        ms_sim = h_sim / (h_sim + c_sim) if (h_sim + c_sim) > 0 else 0
        
        # Write values
        data_cell(ws1, f"A{row}", m_lbl, bold=True, align="left")
        data_cell(ws1, f"B{row}", tariff, num_format="0.0%")
        data_cell(ws1, f"C{row}", p_avg_real, num_format="$#,##0", align="right")
        data_cell(ws1, f"D{row}", p_avg_sim, num_format="$#,##0", align="right")
        data_cell(ws1, f"E{row}", p_var, num_format="0.0%", align="right", color=C_HINO_RED if p_var > 0 else "000000")
        data_cell(ws1, f"F{row}", h_real, num_format="#,##0")
        data_cell(ws1, f"G{row}", h_sim, num_format="#,##0")
        data_cell(ws1, f"H{row}", h_lost, num_format="#,##0", color=C_HINO_RED if h_lost > 0 else "000000", fill="FDEDEC" if h_lost > 0 else None, bold=h_lost > 0)
        data_cell(ws1, f"I{row}", c_real, num_format="#,##0")
        data_cell(ws1, f"J{row}", c_sim, num_format="#,##0")
        data_cell(ws1, f"K{row}", ms_real, num_format="0.0%")
        data_cell(ws1, f"L{row}", ms_sim, num_format="0.0%", fill="E8F8E8" if ms_sim > ms_real else None, bold=ms_sim > ms_real)
        
        data_idx += 1
        
    # Total row
    tot_row = header_row + 1 + data_idx
    ws1.row_dimensions[tot_row].height = 22
    
    # Calculate totals
    tot_h_real = hino_total_real
    tot_h_sim = hino_total_sim
    tot_h_lost = hino_total_lost
    tot_c_real = chev_total_real
    tot_c_sim = chev_total_sim
    
    tot_p_avg_real = (sum((df_hino_sim[f"{m}_VENTAS_REAL"] * df_hino_sim[f"{m}_PRECIO_REAL"]).sum() for m in all_months)) / tot_h_real
    tot_p_avg_sim = (sum((df_hino_sim[f"{m}_VENTAS_SIM"] * df_hino_sim[f"{m}_PRECIO_SIM"]).sum() for m in all_months)) / tot_h_sim
    tot_p_var = (tot_p_avg_real - tot_p_avg_sim) / tot_p_avg_sim
    
    tot_ms_real = tot_h_real / (tot_h_real + tot_c_real)
    tot_ms_sim = tot_h_sim / (tot_h_sim + tot_c_sim)
    
    data_cell(ws1, f"A{tot_row}", "CUMULATIVE TOTAL", fill="EAECEE", bold=True, align="left")
    data_cell(ws1, f"B{tot_row}", "", fill="EAECEE")
    data_cell(ws1, f"C{tot_row}", tot_p_avg_real, fill="EAECEE", bold=True, num_format="$#,##0", align="right")
    data_cell(ws1, f"D{tot_row}", tot_p_avg_sim, fill="EAECEE", bold=True, num_format="$#,##0", align="right")
    data_cell(ws1, f"E{tot_row}", tot_p_var, fill="EAECEE", bold=True, num_format="0.0%", align="right", color=C_HINO_RED if tot_p_var > 0 else "000000")
    data_cell(ws1, f"F{tot_row}", tot_h_real, fill="EAECEE", bold=True, num_format="#,##0")
    data_cell(ws1, f"G{tot_row}", tot_h_sim, fill="EAECEE", bold=True, num_format="#,##0")
    data_cell(ws1, f"H{tot_row}", tot_h_lost, fill="F5B7B1", bold=True, num_format="#,##0", color=C_HINO_RED)
    data_cell(ws1, f"I{tot_row}", tot_c_real, fill="EAECEE", bold=True, num_format="#,##0")
    data_cell(ws1, f"J{tot_row}", tot_c_sim, fill="EAECEE", bold=True, num_format="#,##0")
    data_cell(ws1, f"K{tot_row}", tot_ms_real, fill="EAECEE", bold=True, num_format="0.0%")
    data_cell(ws1, f"L{tot_row}", tot_ms_sim, fill="D5F5E3", bold=True, num_format="0.0%")
    
    # -------------------------------------------------------------
    # TAB 2: DETALLE POR MODELO
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="2. Model Details")
    ws2.sheet_view.showGridLines = True
    
    # Set column widths for sheet 2
    col_widths2 = {
        'A': 22, 'B': 12, 'C': 15, 'D': 15, 'E': 15,
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12,
        'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 12,
        'R': 15, 'S': 18
    }
    for col, w in col_widths2.items():
        ws2.column_dimensions[col].width = w
        
    # Title Block
    ws2.merge_cells("A1:S1")
    merge_title(ws2, "A1", "MODEL-BY-MODEL PRICES AND SALES: REAL VS. SIMULATED (JAN-JUN 2026)", size=12, bold=True)
    ws2.row_dimensions[1].height = 35
    
    # Segment block headers
    # We want a very neat model list
    # Headers
    d_headers = [
        "Hino Model", "Segment", "Jan Price (Base)", "Jun Price (Final)", "Variation %",
        "Jan Real", "Jan Sim", "Feb Real", "Feb Sim", "Mar Real", "Mar Sim",
        "Apr Real", "Apr Sim", "May Real", "May Sim", "Jun Real", "Jun Sim",
        "Lost Sales Volume", "Revenue Loss (USD)"
    ]
    
    ws2.row_dimensions[3].height = 25
    for col_idx, text in enumerate(d_headers):
        cell_ref = f"{get_column_letter(col_idx + 1)}3"
        data_cell(ws2, cell_ref, text, fill=C_DARK_BLUE, bold=True, color=C_WHITE, size=8.5)
        
    # Write Hino models
    for idx, row in df_hino_sim.iterrows():
        r = 4 + idx
        ws2.row_dimensions[r].height = 20
        
        # Calculate values
        m_raw = row['MODEL_RAW']
        segment = row['SEGMENTO']
        p_base = row['2026_01_PRECIO_SIM']
        p_final = row['2026_06_PRECIO_REAL']
        p_var = (p_final - p_base) / p_base if p_base > 0 else 0
        
        lost_total = sum(row[f"{m}_VENTAS_PERDIDAS"] for m in all_months)
        
        # Calculate estimated revenue lost for this specific model
        rev_lost = sum((row[f"{m}_VENTAS_SIM"] * row[f"{m}_PRECIO_SIM"]) - (row[f"{m}_VENTAS_REAL"] * row[f"{m}_PRECIO_REAL"]) for m in all_months)
        
        # Write basic data
        data_cell(ws2, f"A{r}", m_raw, bold=True, align="left")
        data_cell(ws2, f"B{r}", segment)
        data_cell(ws2, f"C{r}", p_base, num_format="$#,##0", align="right")
        data_cell(ws2, f"D{r}", p_final, num_format="$#,##0", align="right")
        data_cell(ws2, f"E{r}", p_var, num_format="0.0%", align="right", color=C_HINO_RED if p_var > 0 else "000000")
        
        # Write monthly sales: Real and Sim
        col_chars = ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]
        month_idx = 0
        for m_db in all_months:
            real_val = row[f"{m_db}_VENTAS_REAL"]
            sim_val = row[f"{m_db}_VENTAS_SIM"]
            
            data_cell(ws2, f"{col_chars[month_idx * 2]}{r}", real_val, num_format="#,##0")
            data_cell(ws2, f"{col_chars[month_idx * 2 + 1]}{r}", sim_val, num_format="#,##0", fill="E8F8E8" if sim_val > real_val else None)
            month_idx += 1
            
        # Write aggregates
        data_cell(ws2, f"R{r}", lost_total, num_format="#,##0", bold=lost_total > 0, fill="FDEDEC" if lost_total > 0 else None, color=C_HINO_RED if lost_total > 0 else "000000")
        data_cell(ws2, f"S{r}", rev_lost, num_format="$#,##0", bold=rev_lost > 0, fill="FDEDEC" if rev_lost > 0 else None, color=C_HINO_RED if rev_lost > 0 else "000000")
        
    # Add Total Row
    tot_row2 = 4 + len(df_hino_sim)
    ws2.row_dimensions[tot_row2].height = 22
    
    data_cell(ws2, f"A{tot_row2}", "TOTAL HINO PORTFOLIO", fill="EAECEE", bold=True, align="left")
    data_cell(ws2, f"B{tot_row2}", "", fill="EAECEE")
    
    # Average base & final price
    avg_p_base = df_hino_sim['2026_01_PRECIO_SIM'].mean()
    avg_p_final = df_hino_sim['2026_06_PRECIO_REAL'].mean()
    avg_p_var = (avg_p_final - avg_p_base) / avg_p_base
    data_cell(ws2, f"C{tot_row2}", avg_p_base, fill="EAECEE", bold=True, num_format="$#,##0", align="right")
    data_cell(ws2, f"D{tot_row2}", avg_p_final, fill="EAECEE", bold=True, num_format="$#,##0", align="right")
    data_cell(ws2, f"E{tot_row2}", avg_p_var, fill="EAECEE", bold=True, num_format="0.0%", align="right")
    
    # Monthly sales sums
    month_col_idx = 0
    for m_db in all_months:
        r_sum = df_hino_sim[f"{m_db}_VENTAS_REAL"].sum()
        s_sum = df_hino_sim[f"{m_db}_VENTAS_SIM"].sum()
        
        data_cell(ws2, f"{col_chars[month_col_idx * 2]}{tot_row2}", r_sum, fill="EAECEE", bold=True, num_format="#,##0")
        data_cell(ws2, f"{col_chars[month_col_idx * 2 + 1]}{tot_row2}", s_sum, fill="D5F5E3", bold=True, num_format="#,##0")
        month_col_idx += 1
        
    # Lost aggregates sums
    tot_lost_agg = df_hino_sim[[f"{m}_VENTAS_PERDIDAS" for m in all_months]].sum().sum()
    tot_rev_lost_agg = sum(sum((df_hino_sim[f"{m}_VENTAS_SIM"] * df_hino_sim[f"{m}_PRECIO_SIM"]) - (df_hino_sim[f"{m}_VENTAS_REAL"] * df_hino_sim[f"{m}_PRECIO_REAL"]) for m in all_months))
    
    data_cell(ws2, f"R{tot_row2}", tot_lost_agg, fill="F5B7B1", bold=True, num_format="#,##0", color=C_HINO_RED)
    data_cell(ws2, f"S{tot_row2}", tot_rev_lost_agg, fill="F5B7B1", bold=True, num_format="$#,##0", color=C_HINO_RED)
    
    # -------------------------------------------------------------
    # TAB 3: COST-PRICE GAP ANALYSIS
    # -------------------------------------------------------------
    df_costs = load_hino_costs()
    ws3 = wb.create_sheet(title="3. Cost-Price Gap Analysis")
    ws3.sheet_view.showGridLines = True
    
    col_widths3 = {
        'A': 22, 'B': 12, 'C': 15, 'D': 12, 'E': 12,
        'F': 15, 'G': 15, 'H': 12, 'I': 12, 'J': 15, 'K': 15, 'L': 18
    }
    for col, w in col_widths3.items():
        ws3.column_dimensions[col].width = w
        
    ws3.merge_cells("A1:L1")
    merge_title(ws3, "A1", "HINO IMPORT COST VS. LIST PRICE MARGIN VARIATION ANALYSIS (JAN-MAY 2026)", size=12, bold=True)
    ws3.row_dimensions[1].height = 35
    
    ws3.merge_cells("A2:L2")
    ws3["A2"] = f"Financial analysis of margin variations comparing percentage cost changes vs. price changes | Parameters: Elasticity={elasticity_light}, June Proj={june_sales_val} | Generated: {datetime.date.today().strftime('%d/%m/%Y')}"
    ws3["A2"].font = Font(name="Arial", size=8, italic=True, color=C_MUTED)
    ws3["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws3["A2"].fill = header_fill("EBF5FB")
    ws3.row_dimensions[2].height = 20
    
    # Merge prices, costs and sales
    df_c_clean = df_costs.copy()
    df_merged_costs = pd.merge(df_pr, df_c_clean, on="MODEL_CLEAN", suffixes=('_PRICE', '_COST'))
    
    df_hino_sim_sales = df_hino_sim.copy()
    hist_months = ["2026_01", "2026_02", "2026_03", "2026_04", "2026_05"]
    df_hino_sim_sales['total_sales_real'] = df_hino_sim_sales[[f"{m}_VENTAS_REAL" for m in hist_months]].sum(axis=1)
    df_tab = pd.merge(df_merged_costs, df_hino_sim_sales[['MODEL_CLEAN', 'total_sales_real']], on="MODEL_CLEAN")
    
    df_tab['price_inc'] = df_tab['MAY_PRICE'] - df_tab['JANUARY_PRICE']
    df_tab['cost_inc'] = df_tab['MAY_COST'] - df_tab['JANUARY_COST']
    df_tab['cost_inc_pct'] = (df_tab['cost_inc'] / df_tab['JANUARY_COST']) * 100
    df_tab['price_inc_pct'] = (df_tab['price_inc'] / df_tab['JANUARY_PRICE']) * 100
    df_tab['margin_var_pp'] = df_tab['price_inc_pct'] - df_tab['cost_inc_pct']
    
    tot_sales_units = df_tab['total_sales_real'].sum()
    
    # KPI Section
    ws3.merge_cells("A4:C4")
    merge_title(ws3, "A4", "COST IMPACT KPI SUMMARY", size=10, bold=True, fill=C_HINO_RED)
    ws3.row_dimensions[4].height = 22
    
    kpis3 = [
        ("Hino Total Sales Volume (units)", tot_sales_units, "#,##0", "Total volume of analyzed Hino models sold (Jan-May 2026)")
    ]
    
    for idx, (kpi_name, kpi_val, kpi_fmt, kpi_desc) in enumerate(kpis3):
        row = 5 + idx
        ws3.row_dimensions[row].height = 20
        ws3.merge_cells(f"A{row}:C{row}")
        data_cell(ws3, f"A{row}", kpi_name, fill=C_DARK_BLUE, bold=True, color=C_WHITE, align="left")
        
        ws3.merge_cells(f"D{row}:E{row}")
        data_cell(ws3, f"D{row}", kpi_val, fill=None, bold=True, color="000000", align="right", num_format=kpi_fmt)
        
        ws3.merge_cells(f"F{row}:L{row}")
        data_cell(ws3, f"F{row}", kpi_desc, align="left", size=8.5, color="555555")
        
    # Section 2: Detailed Table
    start_table_row = 8
    ws3.merge_cells(f"A{start_table_row}:L{start_table_row}")
    merge_title(ws3, f"A{start_table_row}", "MODEL-BY-MODEL COST AND PRICE MARGIN VARIATION BREAKDOWN", size=10, bold=True, fill=C_HINO_RED)
    ws3.row_dimensions[start_table_row].height = 22
    
    headers3 = [
        "Hino Model", "Series", "Sales (units)", "Jan Cost", "May Cost",
        "Cost Inc. ($)", "Cost Inc. (%)", "Jan Price", "May Price", "Price Inc. ($)", "Price Inc. (%)", "Net Margin Var (p.p.)"
    ]
    header_row3 = start_table_row + 1
    ws3.row_dimensions[header_row3].height = 25
    for col_idx, text in enumerate(headers3):
        cell_ref = f"{get_column_letter(col_idx + 1)}{header_row3}"
        data_cell(ws3, cell_ref, text, fill=C_DARK_BLUE, bold=True, color=C_WHITE, size=8.5)
        
    for idx, row_d in df_tab.iterrows():
        r = header_row3 + 1 + idx
        ws3.row_dimensions[r].height = 20
        
        var_val = row_d['margin_var_pp']
        fill_cell = "E8F8F5" if var_val >= 0 else "FDEDEC"  # soft green / soft red
        text_cell = "27AE60" if var_val >= 0 else C_HINO_RED      # green / red
        
        data_cell(ws3, f"A{r}", row_d['MODEL_RAW'], bold=True, align="left")
        data_cell(ws3, f"B{r}", row_d['SERIES'])
        data_cell(ws3, f"C{r}", int(row_d['total_sales_real']), num_format="#,##0")
        data_cell(ws3, f"D{r}", row_d['JANUARY_COST'], num_format="$#,##0", align="right")
        data_cell(ws3, f"E{r}", row_d['MAY_COST'], num_format="$#,##0", align="right")
        data_cell(ws3, f"F{r}", row_d['cost_inc'], num_format="$#,##0", align="right")
        data_cell(ws3, f"G{r}", row_d['cost_inc_pct']/100.0, num_format="+0.0%;-0.0%;0.0%", align="right")
        data_cell(ws3, f"H{r}", row_d['JANUARY_PRICE'], num_format="$#,##0", align="right")
        data_cell(ws3, f"I{r}", row_d['MAY_PRICE'], num_format="$#,##0", align="right")
        data_cell(ws3, f"J{r}", row_d['price_inc'], num_format="$#,##0", align="right")
        data_cell(ws3, f"K{r}", row_d['price_inc_pct']/100.0, num_format="+0.0%;-0.0%;0.0%", align="right")
        data_cell(ws3, f"L{r}", var_val/100.0, num_format="+0.0%;-0.0%;0.0%", align="right", fill=fill_cell, color=text_cell, bold=True)
        
    # Total Row Tab 3
    tot_row3 = header_row3 + 1 + len(df_tab)
    ws3.row_dimensions[tot_row3].height = 22
    
    avg_cost_pct = df_tab['cost_inc_pct'].mean()
    avg_price_pct = df_tab['price_inc_pct'].mean()
    avg_margin_pp = df_tab['margin_var_pp'].mean()
    
    fill_tot = "EAECEE"
    data_cell(ws3, f"A{tot_row3}", "TOTAL PORTFOLIO / AVERAGE", fill=fill_tot, bold=True, align="left")
    data_cell(ws3, f"B{tot_row3}", "", fill=fill_tot)
    data_cell(ws3, f"C{tot_row3}", tot_sales_units, fill=fill_tot, bold=True, num_format="#,##0")
    data_cell(ws3, f"D{tot_row3}", df_tab['JANUARY_COST'].mean(), fill=fill_tot, bold=True, num_format="$#,##0", align="right")
    data_cell(ws3, f"E{tot_row3}", df_tab['MAY_COST'].mean(), fill=fill_tot, bold=True, num_format="$#,##0", align="right")
    data_cell(ws3, f"F{tot_row3}", df_tab['cost_inc'].mean(), fill=fill_tot, bold=True, num_format="$#,##0", align="right")
    data_cell(ws3, f"G{tot_row3}", avg_cost_pct/100.0, fill=fill_tot, bold=True, num_format="+0.0%;-0.0%;0.0%", align="right")
    data_cell(ws3, f"H{tot_row3}", df_tab['JANUARY_PRICE'].mean(), fill=fill_tot, bold=True, num_format="$#,##0", align="right")
    data_cell(ws3, f"I{tot_row3}", df_tab['MAY_PRICE'].mean(), fill=fill_tot, bold=True, num_format="$#,##0", align="right")
    data_cell(ws3, f"J{tot_row3}", df_tab['price_inc'].mean(), fill=fill_tot, bold=True, num_format="$#,##0", align="right")
    data_cell(ws3, f"K{tot_row3}", avg_price_pct/100.0, fill=fill_tot, bold=True, num_format="+0.0%;-0.0%;0.0%", align="right")
    data_cell(ws3, f"L{tot_row3}", avg_margin_pp/100.0, fill=fill_tot, bold=True, num_format="+0.0%;-0.0%;0.0%", align="right", color="27AE60" if avg_margin_pp >= 0 else C_HINO_RED)
    # Series summary in Excel Sheet 3
    start_s_row = tot_row3 + 3
    ws3.merge_cells(f"A{start_s_row}:L{start_s_row}")
    ws3[f"A{start_s_row}"] = "SERIES-LEVEL COST AND PRICE MARGIN VARIATION SUMMARY (WEIGHTED AVERAGES)"
    ws3[f"A{start_s_row}"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    ws3[f"A{start_s_row}"].fill = PatternFill("solid", fgColor=C_HINO_RED)
    ws3[f"A{start_s_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[start_s_row].height = 22
    
    hdr_row_s = start_s_row + 1
    ws3.row_dimensions[hdr_row_s].height = 25
    headers_s = [
        "Hino Series", "", "Sales (units)", "Jan Cost (W)", "May Cost (W)",
        "Cost Inc. ($)", "Cost Inc. (%)", "Jan Price (W)", "May Price (W)", "Price Inc. ($)", "Price Inc. (%)", "Net Margin Var (p.p.)"
    ]
    for col_idx, text in enumerate(headers_s):
        cell_ref = f"{get_column_letter(col_idx + 1)}{hdr_row_s}"
        data_cell(ws3, cell_ref, text, fill=C_DARK_BLUE, bold=True, color="FFFFFF", size=8.5)
    ws3.merge_cells(f"A{hdr_row_s}:B{hdr_row_s}")
    
    # Sort Hino Series by order
    series_order = ["SERIE 300", "SERIE 500", "SERIE 700"]
    series_legend_map = {
        "SERIE 300": "SERIE 300 (LIGHT)",
        "SERIE 500": "SERIE 500 (MEDIUM)",
        "SERIE 700": "SERIE 700 (HEAVY / TRACTO)"
    }
    df_series_groups = sorted(df_tab.groupby('SERIES'), key=lambda g: series_order.index(g[0]) if g[0] in series_order else 99)
    
    for idx_s, (series_name, df_s) in enumerate(df_series_groups):
        r_s = hdr_row_s + 1 + idx_s
        ws3.row_dimensions[r_s].height = 20
        
        sales_vol_s = df_s['total_sales_real'].sum()
        if sales_vol_s == 0:
            continue
            
        w_cost_jan_s = np.average(df_s['JANUARY_COST'], weights=df_s['total_sales_real'])
        w_cost_may_s = np.average(df_s['MAY_COST'], weights=df_s['total_sales_real'])
        w_cost_inc_s = w_cost_may_s - w_cost_jan_s
        w_cost_inc_pct_s = (w_cost_inc_s / w_cost_jan_s) * 100
        
        w_price_jan_s = np.average(df_s['JANUARY_PRICE'], weights=df_s['total_sales_real'])
        w_price_may_s = np.average(df_s['MAY_PRICE'], weights=df_s['total_sales_real'])
        w_price_inc_s = w_price_may_s - w_price_jan_s
        w_price_inc_pct_s = (w_price_inc_s / w_price_jan_s) * 100
        
        w_margin_var_s = w_price_inc_pct_s - w_cost_inc_pct_s
        
        fill_cell_s = "E8F8F5" if w_margin_var_s >= 0 else "FDEDEC"
        text_cell_s = "27AE60" if w_margin_var_s >= 0 else C_HINO_RED
        
        # Write cells
        data_cell(ws3, f"A{r_s}", series_legend_map.get(series_name, series_name), bold=True, align="left")
        data_cell(ws3, f"B{r_s}", "")
        ws3.merge_cells(f"A{r_s}:B{r_s}")
        
        data_cell(ws3, f"C{r_s}", int(sales_vol_s), num_format="#,##0")
        data_cell(ws3, f"D{r_s}", w_cost_jan_s, num_format="$#,##0", align="right")
        data_cell(ws3, f"E{r_s}", w_cost_may_s, num_format="$#,##0", align="right")
        data_cell(ws3, f"F{r_s}", w_cost_inc_s, num_format="$#,##0", align="right")
        data_cell(ws3, f"G{r_s}", w_cost_inc_pct_s/100.0, num_format="+0.0%;-0.0%;0.0%", align="right")
        data_cell(ws3, f"H{r_s}", w_price_jan_s, num_format="$#,##0", align="right")
        data_cell(ws3, f"I{r_s}", w_price_may_s, num_format="$#,##0", align="right")
        data_cell(ws3, f"J{r_s}", w_price_inc_s, num_format="$#,##0", align="right")
        data_cell(ws3, f"K{r_s}", w_price_inc_pct_s/100.0, num_format="+0.0%;-0.0%;0.0%", align="right")
        data_cell(ws3, f"L{r_s}", w_margin_var_s/100.0, num_format="+0.0%;-0.0%;0.0%", align="right", fill=fill_cell_s, color=text_cell_s, bold=True)
        
    # Save file relative to script directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(base_dir, "Reporte_Simulacion_Aranceles_2026.xlsx")
    wb.save(output_path)
    print(f"Excel report saved successfully to: {output_path}")

if __name__ == "__main__":
    generate_excel_report()
